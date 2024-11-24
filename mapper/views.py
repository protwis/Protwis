from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, F, Func, Value, Prefetch
from django.conf import settings
from django.views.generic import TemplateView
from django import forms
from django import template
from protein.models import Protein, ProteinFamily
from common.phylogenetic_tree import PhylogeneticTreeGenerator

import json
from copy import deepcopy
from collections import OrderedDict
import sys

try:
    import importlib.metadata
except ImportError:
    sys.modules['importlib.metadata'] = __import__('importlib_metadata')

import pandas as pd
from sklearn.manifold import TSNE
from sklearn.cluster import KMeans

import openpyxl
import os


class LandingPage(TemplateView):
    template_name = 'mapper/data_mapper_landing.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    @staticmethod
    def keep_by_names(data, names_to_keep):
        data_copy = deepcopy(data)
        if isinstance(data_copy, list):
            # Process each item in the list
            kept_items = [LandingPage.keep_by_names(item, names_to_keep) for item in data_copy]
            # Return only non-None items
            return [item for item in kept_items if item is not None]
        elif isinstance(data_copy, OrderedDict):
            name = data_copy.get('name')
            if name not in names_to_keep.keys():
                if 'children' in data_copy:
                    # Recursively process children
                    data_copy['children'] = LandingPage.keep_by_names(data_copy['children'], names_to_keep)
                    # Remove the 'children' key if it's empty after processing
                    if not data_copy['children']:
                        return None
                else:
                    return None
            else:
                # If the name is in the keep list, update the 'value' field
                data_copy['value'] = names_to_keep[name]['Inner']
                # Process children if present
                if 'children' in data_copy:
                    data_copy['children'] = LandingPage.keep_by_names(data_copy['children'], names_to_keep)
                    if not data_copy['children']:
                        del data_copy['children']
            return data_copy
        return data_copy

    @staticmethod
    def convert_keys(datatree, conversion):
        new_tree = {}
        for key, value in datatree.items():
            # Convert the key using the conversion dictionary
            new_key = conversion.get(key, key)  # Fallback to the original key if no conversion is found
            if isinstance(value, dict):
                # Recursively convert keys of nested dictionaries
                new_tree[new_key] = LandingPage.convert_keys(value, conversion)
            else:
                # If the value is a list (end of the branch), just assign it
                new_tree[new_key] = value
        return new_tree

    @staticmethod
    def reassign_keys(original_dict, key_mappings):
         # Create a new dictionary to store the reassigned keys
         new_dict = {}
         # Iterate over the list of tuples (new_key, old_key)
         for new_key, old_key in key_mappings:
             # Check if the old_key exists in the original dictionary
             if old_key in original_dict:
                 # Assign the value from the original dictionary to the new key
                 new_dict[new_key] = original_dict[old_key]

         return new_dict

    @staticmethod
    def filter_and_extend_dict(big_dict, small_dict):
        def filter_dict(d, keys_set):
            filtered_dict = {}
            for key, value in d.items():
                if isinstance(value, dict):
                    filtered_sub_dict = filter_dict(value, keys_set)
                    if filtered_sub_dict:
                        filtered_dict[key] = filtered_sub_dict
                elif isinstance(value, list):
                    filtered_values = [v for v in value if v in keys_set]
                    if filtered_values:
                        filtered_dict[key] = filtered_values
            return filtered_dict

        def extend_dict(d, additional_values):
            for key, value in d.items():
                if isinstance(value, dict):
                    extend_dict(value, additional_values)
                elif isinstance(value, list):
                    extended_list = []
                    for item in value:
                        if item in additional_values:
                            # Wrap values into a list and add an extra dictionary layer
                            extended_list.append({item: [additional_values[item]['Value1'], additional_values[item]['Value2']]})
                        else:
                            extended_list.append(item)
                    d[key] = extended_list

        # Step 1: Filter the big dictionary
        filtered_dict = filter_dict(big_dict, set(small_dict.keys()))

        # Step 2: Extend the filtered dictionary
        extend_dict(filtered_dict, small_dict)

        return filtered_dict

    @staticmethod
    def filter_dict(d, elements):
        filtered_dict = {}
        for key, value in d.items():
            if isinstance(value, dict):
                filtered_sub_dict = LandingPage.filter_dict(value, elements)
                if filtered_sub_dict:
                    filtered_dict[key] = filtered_sub_dict
            elif isinstance(value, list):
                filtered_values = [v for v in value if v in elements]
                if filtered_values:
                    filtered_dict[key] = filtered_values
        return filtered_dict

    @staticmethod
    def generate_list_plot(listplot): #ADD AN INPUT FILTER DICTIONARY
        # Generate the master dict of protein families

        data = list(listplot.keys())
        names = list(Protein.objects.filter(entry_name__in=data).values_list('name', flat=True))
        # Names conversion dict
        names_dict = Protein.objects.filter(entry_name__in=data).values('entry_name', 'name').order_by('entry_name')
        names_conversion_dict = {item['entry_name']: item['name'] for item in names_dict}
        IUPHAR_to_uniprot_dict = {item['name']: item['entry_name'] for item in names_dict}

        families = ProteinFamily.objects.all()
        datatree = {}
        conversion = {}
        for item in families:
            if len(item.slug) == 3 and item.slug not in datatree.keys():
                datatree[item.slug] = {}
                conversion[item.slug] = item.name
            if len(item.slug) == 7 and item.slug not in datatree[item.slug[:3]].keys():
                datatree[item.slug[:3]][item.slug[:7]] = {}
                conversion[item.slug] = item.name
            if len(item.slug) == 11 and item.slug not in datatree[item.slug[:3]][item.slug[:7]].keys():
                datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]] = []
                conversion[item.slug] = item.name
            if len(item.slug) == 15 and item.slug not in datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]]:
                datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]].append(item.name)

        datatree2 = LandingPage.convert_keys(datatree, conversion)
        datatree2.pop('Parent family', None)
        datatree3 = LandingPage.filter_dict(datatree2, names)
        data_converted = {names_conversion_dict[key]: value for key, value in listplot.items()}
        Data_full = {"NameList": datatree3, "DataPoints": data_converted, "LabelConversionDict":IUPHAR_to_uniprot_dict}
        return Data_full

    @staticmethod
    def generate_GPCRome_data(data):
        #Adjust call to exclude odorants
        all_proteins = Protein.objects.filter(species_id=1, parent_id__isnull=True, accession__isnull=False, family_id__slug__startswith='0').exclude(
                                            family_id__slug__startswith='007'
                                        ).exclude(
                                            family_id__slug__startswith='008'
                                        )

        result_dict = {}
        for prot in all_proteins:
            key = prot.entry_name
            # Initialize the key in result_dict if not already present
            if key not in result_dict:
                result_dict[key] = 'empty'
        
        for key in data:
            if key in result_dict:
                result_dict[key] = data[key]['Value1']
        
        proteins = list(Protein.objects.filter(entry_name__in=result_dict.keys()
            ).values('entry_name', 'name').order_by('entry_name'))

        names_conversion_dict = {item['entry_name']: item['name'] for item in proteins}

        names = list(names_conversion_dict.values())

        IUPHAR_to_uniprot_dict = {item['name']: item['entry_name'] for item in proteins}

        families = ProteinFamily.objects.all()
        datatree = {}
        conversion = {}

        for item in families:
            if len(item.slug) == 3 and item.slug not in datatree.keys():
                datatree[item.slug] = {}
                conversion[item.slug] = item.name
            if len(item.slug) == 7 and item.slug not in datatree[item.slug[:3]].keys():
                datatree[item.slug[:3]][item.slug[:7]] = {}
                conversion[item.slug] = item.name
            if len(item.slug) == 11 and item.slug not in datatree[item.slug[:3]][item.slug[:7]].keys():
                datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]] = []
                conversion[item.slug] = item.name
            if len(item.slug) == 15 and item.slug not in datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]]:
                datatree[item.slug[:3]][item.slug[:7]][item.slug[:11]].append(item.name)

        datatree2 = LandingPage.convert_keys(datatree, conversion)
        datatree2.pop('Parent family', None)
        datatree3 = LandingPage.filter_dict(datatree2, names)
        data_converted = {names_conversion_dict[key]: {'Value1':value} for key, value in result_dict.items()}
        data_full = {"NameList": datatree3, "DataPoints": data_converted, "LabelConversionDict":IUPHAR_to_uniprot_dict}

        return data_full

    @staticmethod
    def generate_tree_plot(input_data): #ADD AN INPUT FILTER DICTIONARY
        ### TREE SECTION
        tree = PhylogeneticTreeGenerator()
        class_a_data = tree.get_tree_data(ProteinFamily.objects.get(name='Class A (Rhodopsin)'))
        class_b1_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class B1 (Secretin)'))
        class_b2_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class B2 (Adhesion)'))
        class_c_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class C (Glutamate)'))
        class_f_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class F (Frizzled)'))
        class_t2_data = tree.get_tree_data(ProteinFamily.objects.get(name__startswith='Class T2 (Taste 2)'))
        ### GETTING NODES
        data_a = class_a_data.get_nodes_dict(None)
        data_b1 = class_b1_data.get_nodes_dict(None)
        data_b2 = class_b2_data.get_nodes_dict(None)
        data_c = class_c_data.get_nodes_dict(None)
        data_f = class_f_data.get_nodes_dict(None)
        data_t2 = class_t2_data.get_nodes_dict(None)
        #Collating everything into a single tree
        general_options = {'depth': 4,
                           'branch_length': {1: 'Class A (Rhodopsin)',
                                             2: 'Alicarboxylic acid',
                                             3: 'Gonadotrophin-releasing hormone',
                                             4: ''},
                           'branch_trunc': 0,
                           'leaf_offset': 30,
                           'anchor': "tree_plot",
                           'label_free': [],
                           'fontSize': {
                                'class': "15px",
                                'ligandtype': "14px",
                                'receptorfamily': "13px",
                                'receptor': "12px"
                            }}
        master_dict = OrderedDict([('name', ''),
                                   ('value', 3000),
                                   ('color', ''),
                                   ('children',[])])
        class_a_dict = OrderedDict([('name', 'Class A (Rhodopsin)'),
                                   ('value', 0),
                                   ('color', 'Red'),
                                   ('children',data_a['children'])])
        class_b1_dict = OrderedDict([('name', 'Class B1 (Secretin)'),
                                   ('value', 0),
                                   ('color', 'Green'),
                                   ('children',data_b1['children'])])
        class_b2_dict = OrderedDict([('name', 'Class B2 (Adhesion)'),
                                  ('value', 0),
                                  ('color', 'Blue'),
                                  ('children',data_b2['children'])])
        class_c_dict = OrderedDict([('name', 'Class C (Glutamate)'),
                                  ('value', 0),
                                  ('color', 'Purple'),
                                  ('children',data_c['children'])])
        class_f_dict = OrderedDict([('name', 'Class F (Frizzled)'),
                                  ('value', 0),
                                  ('color', 'Grey'),
                                  ('children',data_f['children'])])
        class_t2_dict = OrderedDict([('name', 'Class T2 (Taste 2)'),
                                  ('value', 0),
                                  ('color', 'Orange'),
                                  ('children',data_t2['children'])])
        ### APPENDING TO MASTER DICT
        master_dict['children'].append(class_a_dict)
        master_dict['children'].append(class_b1_dict)
        master_dict['children'].append(class_b2_dict)
        master_dict['children'].append(class_c_dict)
        master_dict['children'].append(class_f_dict)
        master_dict['children'].append(class_t2_dict)

        updated_data = {key.replace('_human', ''): value for key, value in input_data.items()}
        circles = {key.replace('_human', '').upper(): {k: v for k, v in value.items()} for key, value in input_data.items()}
        master_dict = LandingPage.keep_by_names(master_dict, updated_data)

        if len(master_dict['children']) == 1:
            master_dict = master_dict['children'][0]
            general_options['depth'] = 3
            general_options['branch_length'] = {1: 'Alicarboxylic acid',
                                             2: 'Gonadotrophin-releasing hormone',
                                             3: ''}
        else:
            pass

        whole_receptors = Protein.objects.prefetch_related("family", "family__parent__parent__parent")
        whole_rec_dict = {}
        for rec in whole_receptors:
            rec_uniprot = rec.entry_short()
            rec_iuphar = rec.family.name.replace("receptor", '').replace("<i>", "").replace("</i>", "").strip()
            if (rec_iuphar[0].isupper()) or (rec_iuphar[0].isdigit()):
                whole_rec_dict[rec_uniprot] = [rec_iuphar]
            else:
                whole_rec_dict[rec_uniprot] = [rec_iuphar.capitalize()]

        return master_dict, general_options, circles, whole_rec_dict

    @staticmethod
    def clustering_test(method, data, data_type):
        # Convert the nested dictionary to a DataFrame
        data = {key.replace('_human', ''): value for key, value in data.items()}
        data_df = pd.DataFrame(data).T
        if 'Value1' in data_df.columns:
            # Example usage
            reduced_df = LandingPage.reduce_and_cluster(data_df, method=method)
            df_merged = pd.merge(reduced_df, data_df['Value2'], left_on='label', right_index=True, how='left')
            df_merged.rename(columns={'Value2': 'fill'}, inplace=True)

            ## add class/ligand_type/receptor_family clusters ##

            # Step 1: Fetch data
            proteins = Protein.objects.filter(
                parent_id__isnull=True, species_id=1
            ).values_list(
                'entry_name',
                "family__parent__parent__parent__name",  # To be renamed as 'Class'
                'family__parent__parent__name',  # To be renamed as 'Ligand type'
                'family__parent__name'  # To be renamed as 'Receptor family'
            )

            # Step 2: Convert to a DataFrame
            proteins_df = pd.DataFrame(list(proteins), columns=['entry_name', 'Class', 'Ligand type', 'Receptor family'])

            # Step 3: Remove '_human' suffix from 'entry_name'
            proteins_df['entry_name'] = proteins_df['entry_name'].str.replace('_human', '')

            # Step 4: Rename 'entry_name' to 'label'
            proteins_df = proteins_df.rename(columns={'entry_name': 'label'})

            # Step 5: Merge with reduced_df on 'label'
            merged_df = pd.merge(df_merged, proteins_df, on='label', how='left')
            # Prepare the data for visualization
            data_json = merged_df.to_json(orient='records')
        else:
            if data_type == 'seq':
                # Get the info of the plot
                full_matrix = LandingPage.generate_full_matrix(method)
                # full_matrix_structure = LandingPage.generate_full_matrix_structure(method)

                # Filter the original fill matrix based on what we use provided
                reduced_input = full_matrix[full_matrix['label'].isin(list(data.keys()))]
                data_df = pd.DataFrame(data).T
                # Merge the dataframes
                df_merged = pd.merge(reduced_input, data_df['Value2'], left_on='label', right_index=True, how='left')
                df_merged.rename(columns={'Value2': 'fill'}, inplace=True)
                # Prepare the data for visualization
                data_json = df_merged.to_json(orient='records')
            elif data_type == 'structure':
                # Get the info of the plot
                full_matrix_structure = LandingPage.generate_full_matrix_structure(method)

                # Filter the original fill matrix based on what we use provided
                reduced_input = full_matrix_structure[full_matrix_structure['label'].isin(list(data.keys()))]
                data_df = pd.DataFrame(data).T
                # Merge the dataframes
                df_merged = pd.merge(reduced_input, data_df['Value2'], left_on='label', right_index=True, how='left')
                df_merged.rename(columns={'Value2': 'fill'}, inplace=True)
                # Prepare the data for visualization
                data_json = df_merged.to_json(orient='records')

        return data_json

    # Generate full similarity matrix for cluster or load existing #
    def generate_full_matrix(method):
        Data_dir = settings.DATA_DIR
        output_file = os.sep.join([Data_dir, 'structure_data', 'HumanGPCRSimilarityAllData_{}.csv'.format(method)])
        # Check if the file exists
        if os.path.exists(output_file):
            # Load the data from the existing file
            merged_df = pd.read_csv(output_file, index_col=0)
        else:
            # Original processing steps
            similarity_matrix_file = os.sep.join([Data_dir, 'structure_data', 'human_gpcr_similarity_data_all_segments.csv'])
            data = pd.read_csv(similarity_matrix_file)
            data = data[['receptor1_entry_name', 'receptor2_entry_name', 'similarity']]
            matrix = data.pivot(index='receptor1_entry_name', columns='receptor2_entry_name', values='similarity')
            matrix.index = matrix.index.str.replace('_human', '', regex=False)
            matrix.columns = matrix.columns.str.replace('_human', '', regex=False)
            matrix = matrix.fillna(100)

            # ---------------------------------------------
            # Step 2: Normalize Similarity Values to [0, 1]
            # ---------------------------------------------

            normalized_matrix = matrix / 100.0

            distance_matrix_df = 1.0 - normalized_matrix

            # Perform reduction and clustering
            reduced_df = LandingPage.reduce_and_cluster(distance_matrix_df, method='tsne')

            reduced_df['label'] = reduced_df['label'].apply(lambda x: x.split('[Human] ')[1] if '[Human] ' in x else x)
            reduced_df['label'] = reduced_df['label'].apply(lambda x: x.split('_human')[0] if '_human' in x else x)

            # add class/ligand_type/receptor_family clusters

            # Step 1: Fetch data
            proteins = Protein.objects.filter(
                parent_id__isnull=True, species_id=1
            ).values_list(
                'entry_name',
                "family__parent__parent__parent__name",  # To be renamed as 'Class'
                'family__parent__parent__name',  # To be renamed as 'Ligand type'
                'family__parent__name'  # To be renamed as 'Receptor family'
            )

            # Step 2: Convert to a DataFrame
            proteins_df = pd.DataFrame(list(proteins), columns=['entry_name', 'Class', 'Ligand type', 'Receptor family'])
            proteins_df.to_excel(os.sep.join([settings.DATA_DIR, 'structure_data', 'All_GPCRs_ligandType_Families.xlsx']),index=False)

            # Step 3: Remove '_human' suffix from 'entry_name'
            proteins_df['entry_name'] = proteins_df['entry_name'].str.replace('_human', '')

            # Step 4: Rename 'entry_name' to 'label'
            proteins_df = proteins_df.rename(columns={'entry_name': 'label'})

            # Step 5: Merge with reduced_df on 'label'
            merged_df = pd.merge(reduced_df, proteins_df, on='label', how='left')

            # Save the reduced DataFrame to a CSV file
            merged_df.to_csv(output_file)

        return merged_df

    @staticmethod
    def reduce_and_cluster(data, method='tsne', n_components=2, n_clusters=5):
        # if method == 'umap':
        #     reducer = umap.UMAP(n_components=n_components, random_state=42)
        if method == 'tsne':
            reducer = TSNE(n_components=n_components, random_state=42)
        # elif method == 'pca':
        #     reducer = PCA(n_components=n_components, random_state=42)
        else:
            raise ValueError("Method should be either 'umap' or 'tsne'")

        reduced_data = reducer.fit_transform(data)

        # Clustering the reduced data
        kmeans = KMeans(n_clusters=n_clusters, random_state=42)
        clusters = kmeans.fit_predict(reduced_data)

        # Prepare the data for D3.js
        df = pd.DataFrame(reduced_data, columns=['x', 'y'])
        df['cluster'] = clusters
        df['label'] = data.index

        return df

    @staticmethod
    def map_to_quartile(value, quartiles):
        if value <= quartiles[0.25]:
            return 10
        elif value <= quartiles[0.5]:
            return 20
        elif value <= quartiles[0.75]:
            return 30
        else:
            return 40

    @staticmethod
    def Label_conversion_info(data):
        # Get list of keys
        Name_list = list(data.keys())
        # Names conversion dict
        names_dict = Protein.objects.filter(entry_name__in=Name_list).values('entry_name', 'name').order_by('entry_name')
        UniProt_to_IUPHAR_converter = {item['entry_name']: item['name'] for item in names_dict}
        IUPHAR_to_UniProt_converter = {item['name']: item['entry_name'] for item in names_dict}
        Label_converter = {'UniProt_to_IUPHAR_converter':UniProt_to_IUPHAR_converter,'IUPHAR_to_UniProt_converter':IUPHAR_to_UniProt_converter}
        return Label_converter

    def post(self, request, *args, **kwargs):
        ### This method handles POST requests for form submission ###

        if request.method == 'POST':

            # Utilize ExcelUploadForm class #
            form = ExcelUploadForm(request.POST,request.FILES)

            # If form is valid #
            if form.is_valid():

                # Get cleaned data #
                file = form.cleaned_data['file']

                # Check if file is .xlsx #
                if not file.name.endswith('.xlsx'):
                    return render(request, self.template_name, {'upload_status': 'Failed','Error_message': "The uploaded file is not an .xlsx file."})
                else:
                    try:
                        workbook = openpyxl.load_workbook(filename=file,read_only=False)
                    except:
                        return render(request, self.template_name, {'upload_status': 'Failed','Error_message': "Unable to load excel file, might be corrupted or not inline with the template file."})

                    if workbook:

                        protein_data = list(Protein.objects.filter(species=1).values_list('entry_name', flat=True).distinct())

                        # Load excel file (workbook) and get sheet names #
                        sheet_names = workbook.sheetnames

                        # Sheets and headers #

                        Sheet_Header_pass_check = [False,False,False,False,False]

                        # Check all sheet names, headers and subheaders (needs to be implemented) #
                        for sheet_name in sheet_names:
                            worksheet = workbook[sheet_name]
                            header_list = [cell.value for cell in worksheet[1]]
                            if sheet_name == 'Tree':
                                Sheet_Header_pass_check[0] = True
                            elif sheet_name == 'Cluster':
                                Sheet_Header_pass_check[1] = True
                            elif sheet_name == 'List':
                                Sheet_Header_pass_check[2] = True
                            elif sheet_name == 'Heatmap':
                                Sheet_Header_pass_check[3] = True
                            elif sheet_name == 'GPCRome':
                                Sheet_Header_pass_check[4] = True
                            else:
                                pass

                        if not all(Sheet_Header_pass_check):
                            # Add addition for the different sheets.
                            return render(request, self.template_name, {'upload_status': 'Failed','Error_message': "The excel file is not structured as the template file. There are incorrect sheet names and data setup."})
                        else:

                            # Init incorrect values #
                            plot_names = ['GPCRome','Tree', 'Cluster', 'List', 'Heatmap']
                            Data = {}
                            Data['Datatypes'] = {}
                            Incorrect_values = {}
                            Heatmap_Label_dict = {}

                            for key in plot_names:
                                Data[key] = {}
                                Incorrect_values[key] = {}

                            Plot_parser = ['Failed','Failed','Failed','Failed','Failed']

                            # For each sheet in the workbook #
                            for sheet_name in sheet_names:

                                # Initialize worksheet #
                                worksheet = workbook[sheet_name]

                                try:
                                    header_list = [cell.value for cell in worksheet[1]]
                                except:
                                    return render(request, self.template_name, {'upload_status': 'Failed','Error_message': "Corrupted excel, sheets not inline with the template file."})

                                # If first sheet is receptor with correct headers #
                                if sheet_name == 'Tree':

                                    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))

                                    # Initialize dictionaries
                                    data_types = [cell.value for cell in worksheet[3]]

                                    Data['Datatypes']['Tree'] = {}
                                    Data['Datatypes']['Tree']['Inner'] = data_types[1]
                                    Data['Datatypes']['Tree']['Outer1'] = data_types[2]
                                    Data['Datatypes']['Tree']['Outer2'] = data_types[3]
                                    Data['Datatypes']['Tree']['Outer3'] = data_types[4]
                                    Data['Datatypes']['Tree']['Outer4'] = data_types[5]
                                    Data['Datatypes']['Tree']['Outer5'] = data_types[6]

                                    for key in header_list:
                                        Incorrect_values[sheet_name][key] = {}

                                    ##########################
                                    # Run through tree sheet #
                                    ##########################
                                    try:

                                        empty_sheet = True  # Initialize the flag
                                        non_empty_count = 0  # Initialize the count for non-empty cells in the first column

                                        # Iterate over rows starting from the second row (min_row=4, excluding the first 3 header rows)
                                        for row in worksheet.iter_rows(min_row=4, values_only=True):
                                            # If the first column is None or empty, ignore the row
                                            if row[0] is None or row[0] == "":
                                                continue
                                            
                                            # Increment the count if the first column has a value
                                            non_empty_count += 1
                                            
                                            # Check only the columns that have headers, skipping the first column
                                            if any(row[i] is not None for i, header in enumerate(header_list[1:], start=1) if header):
                                                empty_sheet = False  # If any other column has a value, set empty_sheet to False

                                        if empty_sheet:
                                            pass
                                        else:
                                            if non_empty_count > 200:
                                                Incorrect_values[sheet_name]['Error'] = "More than 200 Receptors, please provide less than 200 receptors"
                                            else:
                                                # Iterate through rows starting from the third row
                                                for index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=3):
                                                    # Check the "Receptor (Uniprot)" column for correct values
                                                    if row[0] not in protein_data:
                                                        if row[0] is None or row[0] == "":
                                                            pass
                                                        else:
                                                            Incorrect_values[sheet_name][header_list[0]][index] = '"{}" is a invalid entry'.format(row[0])
                                                    else:
                                                        if row[0] not in Data[sheet_name]:
                                                            Data[sheet_name][row[0]] = {}

                                                        # Check each column for data points, boolean values, and float values
                                                        for col_idx, value in enumerate(row):
                                                            if col_idx == 0:
                                                                continue  # Skip the "Receptor (Uniprot)" column and completely empty columns
                                                            elif data_types[col_idx] not in ['Discrete','Continuous']:
                                                                Incorrect_values[sheet_name][header_list[col_idx]] = 'Incorrect datatype'
                                                            else:
                                                                if value is not None:
                                                                    if data_types[col_idx] == 'Discrete':
                                                                        if str(value).lower() not in ['yes', 'no', '1', '0', 'x']:
                                                                            Incorrect_values[sheet_name][header_list[col_idx]][index] = 'Non-Boolean Value'
                                                                        else:
                                                                            if col_idx == 1:
                                                                                Data[sheet_name][row[0]]['Inner'] = 1 if value in ['yes', 'Yes', '1', 'X'] else 0
                                                                            else:
                                                                                Data[sheet_name][row[0]]['Outer{}'.format(col_idx-1)] = 1 if value in ['yes', 'Yes', '1', 'X'] else 0
                                                                    elif data_types[col_idx] == 'Continuous':
                                                                        try:
                                                                            float_value = float(value)
                                                                            if col_idx == 1:
                                                                                Data[sheet_name][row[0]]['Inner'] = float_value
                                                                            else:
                                                                                Data[sheet_name][row[0]]['Outer{}'.format(col_idx-1)] = float_value
                                                                        except ValueError:
                                                                            Incorrect_values[sheet_name][header_list[col_idx]][index] = 'Non-Continuous Value'
                                                                    else:
                                                                        pass
                                                                else:
                                                                    if data_types[col_idx] == 'Discrete' and col_idx == 1:
                                                                        Data[sheet_name][row[0]]['Inner'] = 0


                                        # Check if any values are incorrect #
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data[sheet_name]:
                                            for col_idx in Incorrect_values[sheet_name]:
                                                # Check if there are any assigned index values for this col_idx
                                                if any(Incorrect_values[sheet_name][col_idx].values()):
                                                    # If any index is assigned, set status to 'Partially_success' and break out of the loop
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'

                                        ## Update Plot parser ##
                                        Plot_parser[1] = status
                                    except:
                                        print("Tree failed")

                                ### Cluster ###
                                elif sheet_name == 'Cluster':

                                    # Initialize dictionaries
                                    data_types = [cell.value for cell in worksheet[3]]
                                    for key in header_list:
                                        Incorrect_values[sheet_name][key] = {}
                                    try:

                                        empty_sheet = True  # Initialize the flag

                                        # Iterate over rows starting from the second row (excluding the header row)
                                        for row in worksheet.iter_rows(min_row=4, values_only=True):
                                            # Check only the columns that have headers, skipping the first column
                                            if any(row[i] is not None for i, header in enumerate(header_list[1:], start=1) if header):
                                                empty_sheet = False
                                                break

                                        if empty_sheet:
                                            pass
                                        else:

                                            # Iterate through rows starting from the second row
                                            for index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
                                                # Check the "Receptor (Uniprot)" column for correct values
                                                if row[0] not in protein_data:
                                                    if row[0] is None or row[0] == "":
                                                        pass
                                                    else:
                                                        Incorrect_values[sheet_name][header_list[0]][index] = '"{}" is a invalid entry'.format(row[0])
                                                else:
                                                    if row[0] not in Data[sheet_name]:
                                                        Data[sheet_name][row[0]] = {}

                                                    # Check each column for data points, boolean values, and float values #
                                                    for col_idx, value in enumerate(row):
                                                        if col_idx == 0:
                                                            continue  # Skip the "Receptor (Uniprot)" column and completely empty columns #
                                                        elif data_types[col_idx] not in ['Continuous']:
                                                            Incorrect_values[sheet_name][header_list[col_idx]] = 'Incorrect datatype'
                                                        else:
                                                            if value is not None:
                                                                # Handle the 3 different types of input for Cluster analysis (Boolean, Number, and Text) #
                                                                if data_types[col_idx] in ['Continuous']:
                                                                    try:
                                                                        float_value = float(value)
                                                                        Data[sheet_name][row[0]]['Value{}'.format(col_idx)] = float_value
                                                                    except ValueError:
                                                                        Incorrect_values[sheet_name][header_list[col_idx]][index] = 'Non-Continuous Value'
                                                            else:
                                                                pass

                                        # Check if any values are incorrect #
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data[sheet_name]:
                                            for col_idx in Incorrect_values[sheet_name]:
                                                # Check if there are any assigned index values for this col_idx
                                                if any(Incorrect_values[sheet_name][col_idx].values()):
                                                    # If any index is assigned, set status to 'Partially_success' and break out of the loop
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'

                                        ## Update Plot_parser for Cluster
                                        Plot_parser[2] = status
                                    except:
                                        print("Cluster failed")

                                ### List ###
                                elif sheet_name == 'List':

                                    # Initialize dictionaries
                                    data_types = [cell.value for cell in worksheet[2]]

                                    Data['Datatypes']['Listplot'] = {}
                                    Data['Datatypes']['Listplot']['Col1'] = data_types[2]
                                    Data['Datatypes']['Listplot']['Col2'] = data_types[4]
                                    Data['Datatypes']['Listplot']['Col3'] = data_types[6]
                                    Data['Datatypes']['Listplot']['Col4'] = data_types[8]
                                    for key in header_list:
                                        Incorrect_values[sheet_name][key] = {}
                                    try:
                                        empty_sheet = True  # Initialize the flag

                                        # Iterate over rows starting from the second row (excluding the header row)
                                        for row in worksheet.iter_rows(min_row=3, values_only=True):
                                            # Check only the columns that have headers, skipping the first column
                                            if any(row[i] is not None for i, header in enumerate(header_list[1:], start=1) if header):
                                                empty_sheet = False
                                                break

                                        if empty_sheet:
                                            pass
                                        else:
                                            # Iterate through rows starting from the second row
                                            for index, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
                                                # Check the "Receptor (Uniprot)" column for correct values
                                                if row[0] not in protein_data:
                                                    if row[0] is None or row[0] == "":
                                                        pass
                                                    else:
                                                        Incorrect_values[sheet_name][header_list[0]][index] = '"{}" is a invalid entry'.format(row[0])
                                                else:
                                                    if row[0] not in Data[sheet_name]:
                                                        Data[sheet_name][row[0]] = {}

                                                    # Check each column for data points, boolean values, and float values #
                                                    for col_idx, value in enumerate(row):
                                                        if col_idx == 0:
                                                            continue  # Skip the "Receptor (Uniprot)" column and completely empty columns #
                                                        elif data_types[col_idx] not in ['Boolean','Number','Discrete','Continuous','Text']:
                                                            Incorrect_values[sheet_name][header_list[col_idx]] = 'Incorrect datatype'
                                                        else:
                                                            if value is not None:
                                                                # Handle the 2 different types of input for Cluster analysis (Boolean or Number) #
                                                                if data_types[col_idx] == 'Boolean':
                                                                    if str(value).lower() not in ['yes', 'no', '1', '0']:
                                                                        Incorrect_values[sheet_name][header_list[col_idx]][index] = 'Non-Boolean Value'
                                                                    else:
                                                                        Data[sheet_name][row[0]]['Value{}'.format(col_idx)] = value
                                                                elif data_types[col_idx] == 'Number' or data_types[col_idx] == 'Continuous':
                                                                    try:
                                                                        float_value = float(value)
                                                                        Data[sheet_name][row[0]]['Value{}'.format(col_idx)] = float_value
                                                                    except ValueError:
                                                                        Incorrect_values[sheet_name][header_list[col_idx]][index] = 'Non-Number Value'
                                                                elif data_types[col_idx] == 'Discrete' or data_types[col_idx] == 'Text':
                                                                    Data[sheet_name][row[0]]['Value{}'.format(col_idx)] = value
                                                                else:
                                                                    pass
                                                            else:
                                                                pass

                                        # Check if any values are incorrect #
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data[sheet_name]:
                                            for col_idx in Incorrect_values[sheet_name]:
                                                # Check if there are any assigned index values for this col_idx
                                                if any(Incorrect_values[sheet_name][col_idx].values()):
                                                    # If any index is assigned, set status to 'Partially_success' and break out of the loop
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'

                                        ## Update Plot_parser for Cluster
                                        Plot_parser[3] = status
                                    except:
                                        print("List failed")

                                ###############
                                ### Heatmap ###
                                ###############
                                elif sheet_name == 'Heatmap':

                                    # Initialize dictionaries
                                    data_types = [cell.value for cell in worksheet[2]]
                                    for key in header_list:
                                        Incorrect_values[sheet_name][key] = {}
                                    try:
                                        
                                        empty_sheet = True  # Initialize the flag
                                        non_empty_count = 0  # Initialize the count for non-empty cells in the first column

                                        # Iterate over rows starting from the second row (min_row=4, excluding the first 3 header rows)
                                        for row in worksheet.iter_rows(min_row=3, values_only=True):
                                            # If the first column is None or empty, ignore the row
                                            if row[0] is None or row[0] == "":
                                                continue
                                            
                                            # Increment the count if the first column has a value
                                            non_empty_count += 1
                                            
                                            # Check only the columns that have headers, skipping the first column
                                            if any(row[i] is not None for i, header in enumerate(header_list[1:], start=1) if header):
                                                empty_sheet = False  # If any other column has a value, set empty_sheet to False

                                        if empty_sheet:
                                            pass
                                        else:
                                            if non_empty_count > 50:
                                                Incorrect_values[sheet_name]['Error'] = "More than 50 Receptors, current count: {}, please provide less than 51 receptors".format(non_empty_count)
                                            else:
                                                # create label variables
                                                i = 0
                                                for key in header_list[1:]:
                                                    i += 1
                                                    Heatmap_Label_dict['Value{}'.format(i)] = key
                                                # Iterate through rows starting from the second row
                                                for index, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
                                                    # Check the "Receptor (Uniprot)" column for correct values
                                                    if row[0] not in protein_data:
                                                        if row[0] is None or row[0] == "":
                                                            pass
                                                        else:
                                                            Incorrect_values[sheet_name][header_list[0]][index] = '"{}" is a invalid entry'.format(row[0])
                                                    else:
                                                        if row[0] not in Data[sheet_name]:
                                                            Data[sheet_name][row[0]] = {}

                                                        # Check each column for data points, boolean values, and float values #
                                                        for col_idx, value in enumerate(row):
                                                            if col_idx == 0:
                                                                continue  # Skip the "Receptor (Uniprot)" column and completely empty columns #
                                                            elif data_types[col_idx] not in ['Continuous']:
                                                                Incorrect_values[sheet_name][header_list[col_idx]] = 'Incorrect datatype'
                                                            else:
                                                                if value is not None:
                                                                    # Handle the 1 different types of input for Heatmap (Number) #
                                                                    if data_types[col_idx] == 'Continuous':
                                                                        try:
                                                                            float_value = float(value)
                                                                            Data[sheet_name][row[0]]['Value{}'.format(col_idx)] = float_value
                                                                        except ValueError:
                                                                            Incorrect_values[sheet_name][header_list[col_idx]][index] = 'Non-Continuous Value'
                                                                    else:
                                                                        pass
                                                                else:
                                                                    pass

                                        # Check if any values are incorrect #
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data[sheet_name]:
                                            for col_idx in Incorrect_values[sheet_name]:
                                                # Check if there are any assigned index values for this col_idx
                                                if any(Incorrect_values[sheet_name][col_idx].values()):
                                                    # If any index is assigned, set status to 'Partially_success' and break out of the loop
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'

                                        ## Update Plot_parser for Heatmap
                                        Plot_parser[4] = status
                                    except:
                                        print("Heatmap Failed")

                                ### GPCRome Plot ###
                                elif sheet_name == 'GPCRome':

                                    # Initialize dictionaries
                                    data_types_circle = [cell.value for cell in worksheet[2]]
                                    # Data['Datatypes'] = {}
                                    Data['Datatypes']['GPCRome'] = {}
                                    Data['Datatypes']['GPCRome']['Col1'] = data_types_circle[1]
                                    for key in header_list:
                                        Incorrect_values[sheet_name][key] = {}
                                    try:
                                        empty_sheet = True  # Initialize the flag

                                        # Iterate over rows starting from the second row (excluding the header row)
                                        for row in worksheet.iter_rows(min_row=3, values_only=True):
                                            # Check only the columns that have headers, skipping the first column
                                            if any(row[i] is not None for i, header in enumerate(header_list[1:], start=1) if header):
                                                empty_sheet = False
                                                break

                                        if empty_sheet:
                                            pass
                                        else:
                                            # Iterate through rows starting from the second row
                                            for index, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
                                                # Check the "Receptor (Uniprot)" column for correct values
                                                if row[0] not in protein_data:
                                                    if row[0] is None or row[0] == "":
                                                        pass
                                                    else:
                                                        Incorrect_values[sheet_name][header_list[0]][index] = '"{}" is a invalid entry'.format(row[0])
                                                else:
                                                    if row[0] not in Data[sheet_name]:
                                                        Data[sheet_name][row[0]] = {}

                                                    # Check each column for data points, boolean values, and float values #
                                                    for col_idx, value in enumerate(row):
                                                        if col_idx == 0:
                                                            continue  # Skip the "Receptor (Uniprot)" column and completely empty columns #
                                                        elif data_types_circle[col_idx] not in ['Discrete', 'Continuous']:
                                                            Incorrect_values[sheet_name][header_list[col_idx]] = 'Incorrect datatype'
                                                        else:
                                                            if value is not None:
                                                                # Handle the 2 different types of input for Cluster analysis (Boolean or Number) #
                                                                if data_types_circle[col_idx] == 'Discrete':
                                                                    if str(value).lower() not in ['yes', 'no', '1', '0']:
                                                                        Incorrect_values[sheet_name][header_list[col_idx]][index] = 'Non-Discrete Value'
                                                                    else:
                                                                        Data[sheet_name][row[0]]['Value{}'.format(col_idx)] = value
                                                                elif data_types_circle[col_idx] == 'Continuous':
                                                                    try:
                                                                        float_value = float(value)
                                                                        Data[sheet_name][row[0]]['Value{}'.format(col_idx)] = float_value
                                                                    except ValueError:
                                                                        Incorrect_values[sheet_name][header_list[col_idx]][index] = 'Non-Number Value'
                                                                else:
                                                                    pass
                                                            else:
                                                                pass

                                        # Check if any values are incorrect #
                                        status = 'Success'

                                        if empty_sheet:
                                            status = 'Empty sheet'
                                        elif Data[sheet_name]:
                                            for col_idx in Incorrect_values[sheet_name]:
                                                # Check if there are any assigned index values for this col_idx
                                                if any(Incorrect_values[sheet_name][col_idx].values()):
                                                    # If any index is assigned, set status to 'Partially_success' and break out of the loop
                                                    status = 'Failed'
                                                    break
                                        else:
                                            status = 'Failed'

                                        ## Update Plot_parser for GPCRome
                                        Plot_parser[0] = status
                                    except:
                                        print("GPCRome failed")
                            ## Return all values for plotparser and correctly (or partially) succesful plots ##

                            plot_names = ['GPCRome','Tree', 'Cluster', 'List', 'Heatmap']
                            plot_data = {}
                            plot_incorrect_data = {}

                            for plot_name, plot_status in zip(plot_names, Plot_parser):
                                if plot_status == 'Success':
                                    plot_data[plot_name] = Data[plot_name]
                                elif plot_status == 'Failed':
                                    plot_incorrect_data[plot_name] = Incorrect_values[plot_name]

                            plot_data['Datatypes'] = Data['Datatypes']
                            if 'Heatmap_Label_dict' not in plot_data and Heatmap_Label_dict:
                                plot_data['Heatmap_Label_dict'] = Heatmap_Label_dict

                            plot_data_json = json.dumps(plot_data, indent=4, sort_keys=True) if plot_data else None
                            # plot_incorrect_data_json = json.dumps(plot_incorrect_data, indent=4, sort_keys=True) if plot_incorrect_data else None

                            Plot_parser_json = json.dumps([status == 'Success' for status in Plot_parser])

                            plots_status = [{'status': status, 'plot_name': plot_name} for status, plot_name in zip(Plot_parser, plot_names)]
                            # Rearrange plots in the report #
                            plots_status.sort(key=lambda plot: {'Success': 0, 'Empty sheet': 1, 'Failed': 2}[plot['status']])

                            context = {'upload_status': 'Success',
                                       'report_status': 'Failed',
                                       'Plot_parser':Plot_parser,
                                       'Plot_parser_json':Plot_parser_json,
                                       'plot_names':plot_names,
                                       'plots_status':plots_status}

                            if plot_data:
                                if all(status == 'Success' for status in Plot_parser):
                                    context['report_status'] = 'Success'
                                    context['Data'] = plot_data_json
                                elif 'Success' in Plot_parser and all(status in ['Success', 'Empty sheet'] for status in Plot_parser):
                                    context['report_status'] = 'Partially_success'
                                    context['Data'] = plot_data_json
                            else:
                                context['Data'] = "No Data"

                            if plot_incorrect_data:
                                context['Incorrect_data_json'] = plot_incorrect_data
                            else:
                                context['Incorrect_data_json'] = "No incorrect data"
                            return render(request, self.template_name, context)


                    else:
                        return render(request, self.template_name, {'upload_status': 'Failed','Error_message': "Unable to load excel file, might be corrupted or not inline with the template file."})

            else:
                # Return a 405 Method Not Allowed response if not a POST request
                return render(request, self.template_name, {'upload_status': 'Failed','Error_message': "Not a valid excel file. Please try and use the template excel file."})

#######################
## Excel upload form ##
#######################

class ExcelUploadForm(forms.Form):
    file = forms.FileField()

class plotrender(TemplateView):
    template_name = 'mapper/data_mapper_plotrender.html'

    def post(self, request, *args, **kwargs):
        # Retrieve the sample data from the POST request
        Plot_evaluation_json = request.POST.get('Plot_evaluation')
        Data_json = request.POST.get('Data')
        # If Plot_evaluation_json is not None, parse it as JSON
        if Plot_evaluation_json and Data_json:
            try:
                Plot_evaluation = json.loads(Plot_evaluation_json)
                Data = json.loads(Data_json)
            except json.JSONDecodeError:
                # Handle the case when the JSON data is invalid
                return HttpResponse("Invalid JSON data")
            # Contruct context
            context = {'Plot_evaluation_json': Plot_evaluation}
            context['tree'] = {}
            context['tree_options'] = {}
            context['circles'] = {}
            context['whole_dict'] = {}
            context['cluster_data'] = {}
            context['plot_type'] = {}
            context['heatmap_data'] = {}
            context['listplot_data'] = {}
            if Plot_evaluation:
                # tree #
                if Plot_evaluation[1]:
                    print("Tree success")
                    tree, tree_options, circles, receptors = LandingPage.generate_tree_plot(Data['Tree'])
                    context['tree'] = json.dumps(tree)
                    context['tree_options'] = tree_options
                    context['circles'] = json.dumps(circles)
                    context['whole_dict'] = json.dumps(receptors)
                    context['Tree_datatypes'] = json.dumps(Data['Datatypes'])

                # Cluster analysis #
                if Plot_evaluation[2]:
                    print("Cluster success")
                    output_seq = LandingPage.clustering_test('tsne', Data['Cluster'],'seq')
                    # output_structure = LandingPage.clustering_test('umap', Data['Cluster'],'structure')
                    context['cluster_data_seq'] = output_seq
                    # context['cluster_data_structure'] = output_structure
                    context['plot_type'] = 'Tsne'

                # List plot #
                if Plot_evaluation[3]:
                    print("List success")
                    listplot_data = LandingPage.generate_list_plot(Data['List'])
                    context['listplot_data'] = json.dumps(listplot_data["NameList"])
                    context['listplot_data_variables'] = json.dumps(listplot_data['DataPoints'])
                    context['Label_Conversion'] = json.dumps(listplot_data['LabelConversionDict'])
                    context['listplot_datatypes'] = json.dumps(Data['Datatypes']['Listplot'])

                # Heatmap #
                if Plot_evaluation[4]:
                    print("Heatmap success")
                    label_converter = LandingPage.Label_conversion_info(Data['Heatmap'])
                    context['Label_converter'] = json.dumps(label_converter)
                    context['heatmap_data'] = json.dumps(Data['Heatmap'])
                    context['Heatmap_Label_dict'] = json.dumps(Data['Heatmap_Label_dict'])
                
                # GPCRome #
                if Plot_evaluation[0]:
                    print("GPCRome success")
                    GPCRome_data = LandingPage.generate_GPCRome_data(Data['GPCRome'])
                    context['GPCRome_data'] = json.dumps(GPCRome_data["NameList"])
                    context['GPCRome_data_variables'] = json.dumps(GPCRome_data['DataPoints'])
                    context['GPCRome_Label_Conversion'] = json.dumps(GPCRome_data['LabelConversionDict'])
                    context['GPCRome_datatypes'] = json.dumps(Data['Datatypes'])
                # Handles and determines first active tab #
                first_active_tab = None
                tab_names = ['#tab1', '#tab2', '#tab3', '#tab4','#tab5']

                for i, is_active in enumerate(Plot_evaluation):
                    if is_active:
                        first_active_tab = tab_names[i]
                        break
                context['first_active_tab'] = first_active_tab

            # Return the context dictionary
            return self.render_to_response(context)
        else:
            # Handle the case when Plot_evaluation_json is None
            # This could happen if the form was submitted without the JSON data
            return HttpResponse("Missing sample data")
